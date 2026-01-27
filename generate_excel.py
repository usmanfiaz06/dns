import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────
title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
section_font = Font(name='Arial', size=12, bold=True, color='1B4332')
sub_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
note_font = Font(name='Arial', size=9, italic=True, color='C0392B')
total_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
grand_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
link_font = Font(name='Arial', size=10, color='2E86C1', italic=True)

green_fill = PatternFill(start_color='1B4332', end_color='1B4332', fill_type='solid')
dark_green_fill = PatternFill(start_color='0B2920', end_color='0B2920', fill_type='solid')
header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
light_fill = PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid')
subtotal_fill = PatternFill(start_color='40916C', end_color='40916C', fill_type='solid')
grand_fill = PatternFill(start_color='081C15', end_color='081C15', fill_type='solid')
note_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
alt_fill = PatternFill(start_color='F0F7F4', end_color='F0F7F4', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

sar_format = '#,##0'
pct_format = '0%'

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

def style_header_row(ws, row, cols, fill=header_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_row(ws, row, cols, is_alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_align if c <= 2 else center_align
        cell.border = thin_border
        if is_alt:
            cell.fill = alt_fill

def style_subtotal_row(ws, row, cols, fill=subtotal_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def add_title(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = title_font
    cell.fill = green_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

def add_section(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.alignment = left_align

def add_note(ws, row, text, cols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = note_font
    cell.fill = note_fill
    cell.alignment = left_align


# ════════════════════════════════════════════════════════════
# SHEET 3: FINANCIAL PROPOSAL (BUILD FIRST - other sheets reference it)
# ════════════════════════════════════════════════════════════
ws3 = wb.active
ws3.title = "Financial Proposal"
ws3.sheet_properties.tabColor = "1B4332"

# Column widths
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 20
ws3.column_dimensions['F'].width = 55

r = 1
add_title(ws3, r, "SERA 2026 — FINANCIAL PROPOSAL (DETAILED COST BREAKDOWN)", 6); r += 1
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws3.cell(row=r, column=1, value="Saudi Electricity Regulatory Authority | Currency: SAR | All Prices EXCLUDE 15% VAT | Agency Commission: 15%").font = Font(name='Arial', size=10, italic=True, color='1B4332')
r += 2

# ── EVENTS DATA (Increased pricing for ~5.2M target with 15% commission) ──
# Each event: (name, date, attendance, venue_type, items_list)
# items_list: [(description, unit, qty, unit_price, note), ...]

events = [
    # ═══ Q1 ═══
    ("Q1", None, None, None, None),  # Quarter marker

    ("EVENT 1: Annual Meeting 2026", "February 3", 300, "5-Star Hotel Ballroom", [
        ("Venue Rental (5-star ballroom, full day)", "Full day", 1, 58000, "Ritz-Carlton / Four Seasons / Fairmont"),
        ("Catering – Premium Lunch Buffet", "Per person", 300, 270, "Premium hotel buffet, 10+ items, live stations"),
        ("Catering – 2x Coffee Breaks", "Per person", 300, 65, "Morning + afternoon break, premium selection"),
        ("Stage & Branded Backdrop (8m x 5m)", "Package", 1, 24000, "Custom SERA branded, modular stage"),
        ("LED Video Wall (P2.5, 5m x 3m)", "Rental", 1, 18000, "With dedicated operator"),
        ("Professional Sound System (Line Array)", "Package", 1, 12000, "Concert-grade + 6 wireless mics + monitors"),
        ("Simultaneous Translation (AR/EN)", "Package", 1, 18000, "2 interpreters + booth + headsets for 300"),
        ("Professional Photography (2 photographers)", "Full day", 1, 8000, "500+ edited high-res photos, 3-day delivery"),
        ("Professional Videography (2 cameras)", "Full day", 1, 12000, "5-min highlight + full event recording, 4K"),
        ("Registration System & Name Badges", "Package", 1, 6500, "Digital QR check-in + printed badges + app"),
        ("Printed Materials (agenda, notepad, pen)", "Per person", 300, 45, "Premium branded stationery set"),
        ("Professional MC (Bilingual AR/EN)", "Full event", 1, 10000, "Experienced corporate MC"),
        ("Event Branding & Design", "Package", 1, 15000, "All materials design, backdrop, badges, agenda, social media kit"),
        ("Welcome Lounge & VIP Hospitality", "Package", 1, 12000, "VIP reception area with premium refreshments"),
    ]),

    ("EVENT 2: Founding Day (يوم التأسيس)", "February 22", 300, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("National Décor & Heritage Displays", "Package", 1, 14000, "Founding Day theme, heritage artifacts, timeline display"),
        ("Sound System Rental", "Rental", 1, 5000, "For national anthem + speeches + ambient"),
        ("Commemorative Items (badges, pins, flags)", "Per person", 300, 42, "Custom SERA + Founding Day branding"),
        ("Light Refreshments (Traditional Saudi)", "Per person", 300, 55, "Arabic coffee, dates, Saudi heritage sweets"),
        ("Photography", "3 hours", 1, 4000, "Event documentation"),
        ("Social Media Live Coverage", "Package", 1, 4500, "Real-time posts + stories + reels"),
        ("Event Branding & Design", "Package", 1, 14000, "Heritage-themed signage, displays, social media kit"),
    ]),

    ("EVENT 3: Employee Appreciation Day", "March 2", 300, "5-Star Hotel – Evening Gala", [
        ("Venue Rental (Gala table setup)", "Evening", 1, 52000, "Ballroom with round-table gala setup"),
        ("Catering – Gala Dinner (3-course plated)", "Per person", 300, 330, "Premium plated service + welcome drinks + dessert bar"),
        ("Live Band Entertainment", "Package", 1, 22000, "Professional band, 2.5-hour set"),
        ("Stage, Gala Décor & Lighting Design", "Package", 1, 35000, "Premium gala theme, centerpieces, uplighting, LED features"),
        ("Awards & Crystal Trophies", "Set of 12", 1, 12000, "12 crystal trophies + certificates + frames + award video"),
        ("Photography & Videography (Full team)", "Package", 1, 16000, "2 photo + 2 video, highlight reel + same-day preview"),
        ("Employee Gift Boxes", "Per person", 300, 110, "Premium appreciation gift set with personalized item"),
        ("Professional MC", "Full event", 1, 10000, "Gala host"),
        ("Event Branding & Design", "Package", 1, 15000, "Invitations, backdrop, table cards, menu cards, signage, social media"),
    ]),

    ("EVENT 4: International Women's Day", "March 8", 150, "Boutique Venue – Ladies Only", [
        ("Venue Rental (Ladies section)", "Half day", 1, 24000, "Elegant boutique venue or hotel ladies' space"),
        ("Catering – Ladies' Lunch", "Per person", 150, 230, "Refined plated lunch service + welcome drinks"),
        ("Premium Floral Arrangements", "Package", 1, 12000, "Centerpieces + entrance + stage florals + individual stems"),
        ("Keynote Speaker + Panel (3 speakers)", "Package", 1, 18000, "Female industry leaders"),
        ("Elegant Photo Booth", "Package", 1, 8000, "Premium backdrop, props, instant prints, digital sharing"),
        ("Curated Gift Bags", "Per person", 150, 135, "Skincare, accessories, flowers, personalized card"),
        ("Photography (Female photographer)", "Half day", 1, 5000, "400+ edited photos"),
        ("Event Branding & Design", "Package", 1, 14000, "Invitations, social media kit, signage, digital assets"),
    ]),

    ("EVENT 5: Saudi Flag Day", "March 11", 300, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("National Flag Display & Décor", "Package", 1, 12000, "Large flag (5m+), national colors throughout, entrance display"),
        ("Sound System Rental", "Rental", 1, 5000, "For national anthem + speeches + ambient"),
        ("Commemorative Items (pins, badges, flags)", "Per person", 300, 38, "Custom SERA + flag branding, premium quality"),
        ("Light Refreshments", "Per person", 300, 50, "Arabic coffee (dallah), premium dates, Saudi sweets"),
        ("Photography", "2 hours", 1, 3500, "Event documentation"),
        ("Social Media Live Coverage", "Package", 1, 4500, "Real-time posts + stories + reels"),
        ("Event Branding & Design", "Package", 1, 14000, "National day signage, displays, social media, digital assets"),
    ]),

    ("EVENT 6: Ramadan Iftar", "Mid-March", 300, "5-Star Hotel Ramadan Tent", [
        ("Iftar Buffet (5-Star Hotel)", "Per person", 300, 450, "[NOTE] Ramadan premium pricing – 20-30% above regular season"),
        ("VIP Seating Upgrade (Senior Leadership)", "Per person", 40, 220, "[NOTE] Assumed 13% VIP guests – confirm with SERA"),
        ("Traditional Entertainment (Oud + Singer)", "Package", 1, 15000, "Oud player + traditional vocalist + Tawashih"),
        ("Ramadan Décor Enhancement", "Package", 1, 12000, "Lanterns, crescents, fabric draping, ambient lighting, entrance"),
        ("Premium Ramadan Gift Boxes", "Per person", 300, 120, "Premium dates, tasbih, Ramadan book, oud, prayer rug"),
        ("Photography & Videography", "Package", 1, 10000, "Evening iftar coverage + highlight reel"),
        ("Event Branding & Design", "Package", 1, 14000, "Ramadan-themed invitations, signage, social media kit, digital assets"),
    ]),

    ("EVENT 7: Mother's Day", "March 21", 130, "Elegant Restaurant", [
        ("Venue + Lunch (Private dining)", "Per person", 130, 280, "Private restaurant section, premium setting"),
        ("Individual Flower Bouquet", "Per person", 130, 75, "Premium wrapped bouquet per attendee"),
        ("Special Gift for Mothers", "Per person", 65, 185, "[NOTE] Assumed 50% are mothers – adjust based on SERA data"),
        ("Light Entertainment / Performance", "Package", 1, 10000, "Musical performance + singer"),
        ("Photography", "3 hours", 1, 4500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Invitations, signage, cards, social media kit, digital assets"),
    ]),

    ("EVENT 8: Saudi Green Initiative (المبادرة الخضراء)", "March 27", 250, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Green Initiative Exhibition Booths", "Package", 1, 15000, "Interactive sustainability displays, SERA green projects"),
        ("Environmental Expert Speaker", "Fee", 1, 15000, "Sustainability / clean energy expert"),
        ("Awareness Materials & Eco Giveaways", "Per person", 250, 32, "Eco-friendly branded items, seed kits, reusable bags"),
        ("Light Refreshments (Organic/Healthy)", "Per person", 250, 45, "Organic options, minimal-waste serving"),
        ("Photography", "2 hours", 1, 3500, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Green-themed signage, displays, social media kit"),
    ]),

    ("EVENT 9: Ehsan Philanthropy Campaign (حملة إحسان)", "March", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Campaign Materials & Setup", "Package", 1, 10000, "Ehsan platform displays, donation stations, banners"),
        ("Awareness Display & Interactive Booth", "Package", 1, 8000, "Ehsan success stories, impact display, QR donation"),
        ("Ehsan-Branded Giveaways", "Per person", 200, 35, "Branded awareness items, charity wristbands"),
        ("Light Refreshments", "Per person", 200, 40, "Coffee, snacks, sweets"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Campaign signage, displays, social media kit"),
    ]),

    # ═══ Q2 ═══
    ("Q2", None, None, None, None),

    ("EVENT 10: Eid Al-Fitr Celebration", "Early April", 300, "5-Star Hotel – Grand Gala", [
        ("Venue Rental (Grand Ballroom)", "Evening", 1, 58000, "Premium 5-star hotel, premium setup"),
        ("Catering – Eid Gala Dinner", "Per person", 300, 360, "Eid special menu + live cooking stations + dessert bar"),
        ("Stage & Full Eid Décor (Gold/White theme)", "Package", 1, 42000, "Premium festive Eid decoration + entrance arch + lighting"),
        ("Entertainment – Ardah + Live Music", "Package", 1, 30000, "Traditional Ardah troupe + modern band"),
        ("LED Video Wall & Full AV Production", "Package", 1, 22000, "Full stage production, lighting design, effects"),
        ("Photography & Videography (Full team)", "Package", 1, 18000, "Complete coverage, same-day preview, highlight reel"),
        ("Eid Employee Gift Boxes", "Per person", 300, 155, "Premium Eid gift box with personalized item"),
        ("Children's Corner", "Package", 1, 15000, "[NOTE] Optional – only if families invited. Confirm with SERA"),
        ("Professional MC (Bilingual)", "Full event", 1, 12000, "Eid gala host"),
        ("Event Branding & Design", "Package", 1, 16000, "Eid-themed invites, backdrop, social media kit, signage, digital assets"),
    ]),

    ("EVENT 11: Creativity & Innovation Day", "April 21", 200, "Conference Center", [
        ("Venue Rental (Conference + Exhibition)", "Full day", 1, 38000, "Conference center or business hotel with exhibition space"),
        ("Catering (Lunch + 2 Coffee Breaks)", "Per person", 200, 200, "Full-day premium package"),
        ("Innovation Showcase Booths (10 depts)", "Per booth", 10, 5500, "Custom booth design + displays + lighting + interactive"),
        ("Main Stage & Presentation Setup", "Package", 1, 18000, "Stage, LED screen, sound, podium"),
        ("Innovation Awards & Trophies", "Set of 7", 1, 10000, "7 innovation category crystal awards + certificates"),
        ("Keynote Speaker (Innovation Expert)", "Fee", 1, 25000, "[NOTE] Industry expert – price varies by profile"),
        ("Digital Voting / Interactive System", "Rental", 1, 8000, "Audience votes + live results display on screen"),
        ("Photography & Videography", "Package", 1, 12000, "Full day coverage + innovation showcase stories"),
        ("Event Branding & Design", "Package", 1, 15000, "Booth templates, signage, programs, badges, digital content"),
    ]),

    ("EVENT 12: International Tea Day", "May 21", 250, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no cost"),
        ("Professional Tea Stations (3 types)", "Package", 1, 20000, "Arabic, English high tea, Asian ceremony with experts"),
        ("Tea Snacks & Gourmet Pastries", "Per person", 250, 50, "Curated tea pairings, premium pastries"),
        ("Tea-Themed Décor & Setup", "Package", 1, 8000, "Full theming of common area"),
        ("Educational Materials & Displays", "Package", 1, 4000, "Tea culture info + branding + takeaway samples"),
        ("Photography", "2 hours", 1, 3500, "Event coverage"),
    ]),

    ("EVENT 13: Joy Clothing Drive (كسوة فرح)", "May", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Collection Booths & Sorting Setup", "Package", 1, 8000, "Collection points, sorting tables, storage boxes"),
        ("Awareness Campaign Materials", "Package", 1, 8000, "Posters, emails, digital screens, internal comms"),
        ("Volunteer Coordination", "Package", 1, 5000, "Volunteer management, logistics, distribution plan"),
        ("Sorting & Distribution Supplies", "Package", 1, 6000, "Bags, labels, transport boxes, cleaning supplies"),
        ("Photography & Videography", "Package", 1, 5000, "CSR documentation + highlight video"),
        ("Event Branding & Design", "Package", 1, 12000, "Campaign signage, volunteer badges, social media kit"),
    ]),

    ("EVENT 14: World No Tobacco Day (اليوم العالمي لمكافحة التدخين)", "May 31", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Health Expert Speaker", "Fee", 1, 12000, "Pulmonologist or smoking cessation specialist"),
        ("Awareness Materials & Displays", "Package", 1, 10000, "Health impact displays, lung models, interactive demos"),
        ("Nicotine Dependency Assessment Station", "Package", 1, 8000, "CO breath test, health screening station"),
        ("Light Healthy Refreshments", "Per person", 200, 40, "Healthy snacks, fruits, juices, herbal teas"),
        ("Awareness Giveaways", "Per person", 200, 28, "Anti-smoking wristbands, stress balls, info booklets"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Health-themed signage, displays, social media kit"),
    ]),

    ("EVENT 15: Blood Donation Day", "June 14", 120, "SERA HQ", [
        ("Venue Setup", "SERA HQ", 1, 0, "Internal + Red Crescent coordination"),
        ("Pre-event Awareness Campaign", "Package", 1, 10000, "Posters, emails, digital screens, social, internal comms"),
        ("Donor Refreshments", "Per donor", 120, 60, "Juice, snacks, recovery food, premium options"),
        ("Donor Appreciation Gifts", "Per donor", 120, 55, "Branded thank-you gift set"),
        ("Certificates of Appreciation", "Per donor", 120, 25, "Premium branded certificates"),
        ("Photography", "Full day", 1, 3500, "Documentation"),
        ("Red Crescent Coordination", "Partnership", 1, 0, "[NOTE] They provide equipment & staff – confirm partnership"),
    ]),

    ("EVENT 16: Father's Day", "June 15", 160, "Event Space / Restaurant", [
        ("Venue + Lunch/Dinner", "Per person", 160, 260, "Private premium venue booking"),
        ("Father's Day Gifts", "Per person", 80, 160, "[NOTE] Assumed 50% are fathers – adjust per actual"),
        ("Team Activities / Entertainment", "Package", 1, 15000, "Team competitions, entertainment, prizes"),
        ("Photography", "3 hours", 1, 4500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Invitations, signage, décor, social media kit, digital assets"),
    ]),

    ("EVENT 17: Eid Al-Adha Celebration", "Mid-June", 300, "5-Star Hotel – Grand Gala", [
        ("Venue Rental (Grand Ballroom)", "Evening", 1, 58000, "Premium 5-star hotel"),
        ("Catering – Eid Al-Adha Gala Dinner", "Per person", 300, 360, "Traditional lamb feast + premium menu + live stations"),
        ("Stage & Full Eid Décor (Arabian theme)", "Package", 1, 42000, "Traditional Arabian Eid Al-Adha theme + entrance"),
        ("Entertainment – Ardah + Live Music", "Package", 1, 30000, "Traditional + modern acts"),
        ("LED Video Wall & Full AV Production", "Package", 1, 22000, "Full production setup + lighting design"),
        ("Photography & Videography (Full team)", "Package", 1, 18000, "Complete coverage + highlight reel"),
        ("Eid Employee Gift Boxes", "Per person", 300, 155, "Premium gift box with personalized item"),
        ("Professional MC", "Full event", 1, 12000, "Gala host"),
        ("Event Branding & Design", "Package", 1, 16000, "Eid-themed invites, backdrop, social media kit, signage, digital assets"),
    ]),

    # ═══ Q3 ═══
    ("Q3", None, None, None, None),

    ("EVENT 18: Summer SERA (صيف سيرا)", "July 20", 500, "Entertainment Venue – Evening", [
        ("Venue Rental (Entertainment complex)", "Evening", 1, 72000, "[NOTE] MUST be evening post-sunset – July heat exceeds 45°C"),
        ("Catering – Food Festival Style", "Per person", 500, 145, "10+ themed food stations, live cooking"),
        ("Beverages & Refreshments Station", "Per person", 500, 38, "Water, fresh juices, soft drinks, ice cream bar"),
        ("Main Stage Entertainment (Live show)", "Package", 1, 38000, "Headline act + band + performer"),
        ("Children's Entertainment Zone", "Package", 1, 28000, "Games, characters, face painting, bouncy castles, VR"),
        ("Game Booths & Activity Stations", "Per booth", 10, 5000, "10 interactive branded stations with prizes"),
        ("Prizes, Raffle & Giveaways", "Package", 1, 25000, "Competition prizes + grand raffle + spot prizes"),
        ("Summer-Themed Décor (Multi-zone)", "Package", 1, 30000, "Full venue theming across all zones"),
        ("Outdoor Concert Sound & Lighting", "Package", 1, 35000, "Concert-grade outdoor production + effects"),
        ("Photography & Videography (Multi-zone)", "Package", 1, 22000, "3 photographers + 2 videographers + drone"),
        ("Family Gift Bags", "Per family", 250, 90, "Summer activity pack with branded items"),
    ]),

    ("EVENT 19: SERA Winter (شتوية سيرا)", "August 19", 300, "Indoor Entertainment Venue", [
        ("Venue Rental (Indoor entertainment complex)", "Evening", 1, 38000, "Air-conditioned indoor venue – August heat"),
        ("Catering – Food Stations", "Per person", 300, 160, "Themed food stations, live cooking, dessert bar"),
        ("Winter-Themed Décor & Setup", "Package", 1, 18000, "Winter wonderland theme, snow effects, ambient lighting"),
        ("Entertainment Program", "Package", 1, 20000, "Live performance, games, competitions"),
        ("Photo Booth (Winter Theme)", "Package", 1, 8000, "Themed backdrop, props, instant prints"),
        ("Photography & Videography", "Package", 1, 12000, "Full coverage + highlight reel"),
        ("Event Branding & Design", "Package", 1, 14000, "Winter-themed invitations, signage, social media kit"),
    ]),

    ("EVENT 20: World First Aid Day (اليوم العالمي للإسعافات الأولية)", "September 13", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("First Aid Training Workshop", "Package", 1, 15000, "Certified trainer, CPR + first aid techniques"),
        ("CPR Mannequins & Training Equipment", "Rental", 1, 8000, "Professional training dummies + AED trainers"),
        ("First Aid Kits (Giveaway)", "Per person", 200, 45, "Branded personal first aid kits"),
        ("Expert Paramedic Speaker", "Fee", 1, 10000, "Emergency medicine specialist"),
        ("Light Refreshments", "Per person", 200, 40, "Coffee, snacks, water"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Health-themed signage, certificates, social media kit"),
    ]),

    ("EVENT 21: تزود Skills Development (تزوّد)", "September", 200, "Conference / Training Venue", [
        ("Venue Rental (Conference / training room)", "Full day", 1, 25000, "Professional training environment"),
        ("Expert Trainers (2 sessions)", "Fee", 2, 15000, "Industry experts for skills development"),
        ("Training Materials & Workbooks", "Per person", 200, 50, "Branded workbooks, handouts, tools"),
        ("Catering (Lunch + Coffee Break)", "Per person", 200, 150, "Full-day premium catering package"),
        ("Completion Certificates", "Per person", 200, 25, "Premium branded certificates"),
        ("Photography", "Full day", 1, 3500, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Program signage, materials, certificates, social media kit"),
    ]),

    ("EVENT 22: World Alzheimer's Day", "September 21", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal"),
        ("Awareness Materials & Booth Setup", "Package", 1, 10000, "Purple theme + interactive info booth + displays"),
        ("Expert Healthcare Speaker", "Fee", 1, 15000, "Neurologist or awareness specialist"),
        ("Light Refreshments", "Per person", 200, 45, "Coffee, healthy snacks, fruits"),
        ("Purple Ribbons & Awareness Items", "Per person", 200, 20, "Branded awareness items + informational booklet"),
        ("Photography", "2 hours", 1, 3500, "Event coverage"),
    ]),

    ("EVENT 23: وش دورنا – What's Our Role (وش دورنا)", "September", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Workshop Facilitation", "Package", 1, 15000, "Professional facilitator, interactive sessions"),
        ("Interactive Materials & Displays", "Package", 1, 10000, "Role clarity exercises, team activities, displays"),
        ("Light Refreshments", "Per person", 200, 45, "Coffee, snacks, lunch boxes"),
        ("Awareness Materials", "Per person", 200, 28, "Branded booklets, role guides, takeaways"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Workshop signage, materials, social media kit"),
    ]),

    ("EVENT 24: Saudi National Day (96th)", "September 23", 300, "Premium Venue – Evening", [
        ("Venue Rental (Premium + outdoor area)", "Evening", 1, 65000, "[NOTE] Book 3-4 months ahead – extreme demand period"),
        ("Catering – Saudi Traditional Feast", "Per person", 300, 320, "Traditional Saudi dishes + modern, coffee ceremony, live stations"),
        ("Stage & Premium National Décor (Green)", "Package", 1, 48000, "Flags, national symbols, green theme, immersive design"),
        ("Ardah Performance Troupe", "Package", 1, 24000, "Authentic Saudi traditional dance + sword dance"),
        ("Saudi Cultural Entertainment", "Package", 1, 18000, "Additional cultural acts + Saudi music + poet"),
        ("LED Video Wall & AV Production", "Package", 1, 24000, "National content, anthem, videos on LED, special effects"),
        ("Photography & Videography (Full team)", "Package", 1, 20000, "Premium coverage + drone + same-day preview"),
        ("National Day Themed Gifts", "Per person", 300, 125, "Patriotic premium gift items with SERA branding"),
        ("Green Lighting & Special Effects", "Package", 1, 20000, "Building/venue green wash + laser + effects"),
        ("Professional MC", "Full event", 1, 12000, "National Day host"),
        ("Event Branding & Design", "Package", 1, 16000, "Invitations, signage, social media kit, print, digital assets"),
        ("Fireworks / Confetti Finale", "Package", 1, 15000, "[NOTE] Indoor confetti/CO2 if fireworks not permitted"),
    ]),

    ("EVENT 25: Environment Week (أسبوع البيئة)", "Q3 TBD", 250, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Environment Exhibition & Booths (5 days)", "Package", 1, 20000, "Sustainability displays, recycling demos, green projects"),
        ("Expert Speakers (3 sessions)", "Fee", 3, 8000, "Environment & sustainability experts across the week"),
        ("Awareness Materials & Eco Giveaways", "Per person", 250, 30, "Eco-friendly items, seed kits, reusable products"),
        ("Green Activities & Workshops", "Package", 1, 15000, "Tree planting, upcycling workshop, waste reduction demos"),
        ("Light Refreshments (5 days)", "Per person", 250, 40, "Daily healthy refreshments, minimal-waste serving"),
        ("Photography (Full week)", "Package", 1, 10000, "Daily coverage + week highlight reel"),
        ("Event Branding & Design", "Package", 1, 14000, "Environment-themed signage, booth designs, social media kit"),
    ]),

    # ═══ Q4 ═══
    ("Q4", None, None, None, None),

    ("EVENT 26: International Coffee Day", "October 1", 250, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal"),
        ("Professional Barista Station", "Package", 1, 14000, "Specialty coffee equipment + 2 baristas"),
        ("Arabic Coffee Ceremony (Dallah)", "Package", 1, 8000, "Traditional service with premium dates + Saudiya coffee"),
        ("International Coffee + Pastries", "Per person", 250, 55, "Various coffees + gourmet pastries + chocolate"),
        ("Coffee-Themed Décor & Setup", "Package", 1, 7000, "Full theming + bean displays"),
        ("Photography", "2 hours", 1, 3500, "Event coverage"),
    ]),

    ("EVENT 27: Breast Cancer Awareness Day (اليوم العالمي لسرطان الثدي)", "October", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Pink Décor & Awareness Displays", "Package", 1, 12000, "Pink ribbon theme, survivor stories, info panels"),
        ("Healthcare Expert Speaker", "Fee", 1, 12000, "Oncologist or breast cancer awareness specialist"),
        ("Screening Awareness Materials", "Per person", 200, 25, "Self-exam guides, awareness booklets, info cards"),
        ("Pink Ribbons & Awareness Items", "Per person", 200, 30, "Pink ribbon pins, branded awareness items"),
        ("Light Refreshments", "Per person", 200, 40, "Coffee, healthy snacks, pink-themed treats"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Pink-themed signage, displays, social media kit"),
    ]),

    ("EVENT 28: Cybersecurity Exhibition", "October TBD", 250, "SERA HQ / Conference", [
        ("Venue Setup", "SERA HQ", 1, 12000, "Internal setup + additional infrastructure + networking"),
        ("Interactive Exhibition Booths", "Per booth", 6, 14000, "6 interactive cybersecurity demo stations with screens"),
        ("Cybersecurity Expert Speakers", "Fee", 2, 15000, "2 industry-leading experts"),
        ("Interactive Simulations & Demos", "Package", 1, 22000, "Phishing demos, password games, VR, AI threats"),
        ("Awareness Materials & Branded Giveaways", "Per person", 250, 60, "Secure USB, materials, guides, branded items"),
        ("Catering (Full day)", "Per person", 250, 165, "Lunch + 2 premium coffee breaks"),
        ("Photography & Videography", "Package", 1, 12000, "Full day coverage + exhibition walkthrough video"),
        ("Event Branding & Design", "Package", 1, 15000, "Booth designs, signage, materials, badges, digital content"),
    ]),

    ("EVENT 29: World Mental Health Day", "October 10", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal"),
        ("Wellness Expert / Psychologist", "Fee", 1, 18000, "Mental health keynote speaker"),
        ("Stress Management Workshop", "Package", 1, 15000, "Interactive workshop + materials + exercises"),
        ("Relaxation Zone Setup", "Package", 1, 12000, "Calm space, massage chairs, aromatherapy, sound therapy"),
        ("Healthy Refreshments", "Per person", 200, 50, "Herbal teas, fruits, healthy snacks, smoothies"),
        ("Wellness Giveaways", "Per person", 200, 45, "Journals, essential oils, wellness kits, meditation app"),
        ("Photography", "Half day", 1, 3500, "Event coverage"),
    ]),

    ("EVENT 30: World Savings Day (اليوم العالمي للادخار)", "October 31", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Financial Expert Speaker", "Fee", 1, 12000, "Financial planning / savings specialist"),
        ("Awareness Materials & Displays", "Package", 1, 8000, "Savings tips displays, financial literacy info panels"),
        ("Financial Literacy Booklets", "Per person", 200, 25, "Branded savings guides, budget planners"),
        ("Light Refreshments", "Per person", 200, 40, "Coffee, snacks"),
        ("Savings Pledge Wall", "Package", 1, 5000, "Interactive pledge wall + goal-setting station"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Financial-themed signage, materials, social media kit"),
    ]),

    ("EVENT 31: Quality Day", "November 10", 180, "Business Hotel Conference", [
        ("Venue (Conference room)", "Half day", 1, 25000, "Professional conference setting"),
        ("Catering (Lunch + 1 break)", "Per person", 180, 145, "Half-day premium package"),
        ("Quality Awards & Recognition", "Set of 7", 1, 10000, "7 quality awards + certificates + frames"),
        ("Stage & Presentation Setup", "Package", 1, 10000, "Stage, screen, sound, podium, backdrop"),
        ("Quality Expert Keynote Speaker", "Fee", 1, 12000, "Quality/excellence expert"),
        ("Photography", "Half day", 1, 3500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Signage, certificates, backdrop, materials, social media, digital"),
    ]),

    ("EVENT 32: Flu Vaccination Campaign (حملة التطعيم ضد الإنفلونزا)", "November", 250, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Pre-Campaign Awareness Materials", "Package", 1, 8000, "Posters, emails, digital screens, internal comms"),
        ("Medical Staff & Coordination", "Package", 1, 12000, "[NOTE] Nurses + logistics – coordinate with health provider"),
        ("Vaccination Station Setup & Supplies", "Package", 1, 10000, "Medical setup, privacy screens, waiting area"),
        ("Post-Vaccination Refreshments", "Per person", 250, 30, "Juice, snacks, recovery food"),
        ("Appreciation Items", "Per person", 250, 25, "Thank-you gifts for participants"),
        ("Photography", "Full day", 1, 3000, "Health campaign documentation"),
        ("Event Branding & Design", "Package", 1, 12000, "Health-themed signage, awareness displays, social media kit"),
    ]),

    ("EVENT 33: World Diabetes Day (اليوم العالمي للسكري)", "November 14", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Health Expert Speaker", "Fee", 1, 12000, "Endocrinologist or diabetes specialist"),
        ("Diabetes Screening Station", "Package", 1, 10000, "Blood sugar testing, health risk assessment"),
        ("Awareness Materials & Displays", "Per person", 200, 25, "Diabetes info booklets, diet guides, health cards"),
        ("Healthy Refreshments", "Per person", 200, 45, "Sugar-free options, fruits, healthy snacks"),
        ("Blue Circle Awareness Items", "Per person", 200, 30, "Blue circle pins, branded awareness items"),
        ("Photography", "2 hours", 1, 3000, "Event documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Blue-themed signage, health displays, social media kit"),
    ]),

    ("EVENT 34: International Men's Day", "November 19", 200, "Event Space", [
        ("Venue + Lunch/Dinner", "Per person", 200, 210, "Private premium event space"),
        ("Team Activities & Competitions", "Package", 1, 20000, "Team-building, sports competitions, prizes"),
        ("Recognition Gifts", "Per person", 200, 65, "Men's day premium gift"),
        ("Photography", "3 hours", 1, 4500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Invitations, signage, décor, social media kit, digital assets"),
    ]),

    ("EVENT 35: Universal Children's Day", "November 20", 200, "Family Entertainment Venue", [
        ("Venue (Family-friendly)", "4 hours", 1, 30000, "Premium entertainment venue"),
        ("Catering (Kids + Adults)", "Per person", 200, 120, "Kid-friendly + adult menu + dessert station"),
        ("Children's Professional Entertainment", "Package", 1, 25000, "Characters, clown, games, face paint, magic, VR"),
        ("Gifts for Children", "Per child", 100, 80, "Premium toys, activity kits, branded items"),
        ("Photography & Videography", "Package", 1, 6000, "Family event coverage + fun video"),
        ("Event Branding & Design", "Package", 1, 14000, "Kid-friendly branded materials, photo spots, social media, digital"),
    ]),

    ("EVENT 36: Volunteer Day", "December 5", 80, "Community Location", [
        ("Transportation (Buses)", "Per bus", 2, 3500, "SERA HQ to community location + return"),
        ("Branded Volunteer T-Shirts + Caps", "Per person", 80, 70, "SERA branded volunteer kit"),
        ("Activity Materials & Supplies", "Package", 1, 10000, "Supplies for volunteer work + tools"),
        ("Volunteer Lunch Boxes + Water", "Per person", 80, 70, "Premium boxed lunches + snacks"),
        ("Photography & Videography", "Package", 1, 7000, "Documentation + CSR highlight video"),
        ("Appreciation Certificates & Medals", "Per person", 80, 25, "Premium volunteer recognition"),
    ]),

    ("EVENT 37: Anti-Corruption Day", "December 9", 200, "SERA HQ / Conference", [
        ("Venue", "SERA HQ", 1, 0, "Internal"),
        ("Ethics Expert / Nazaha Speaker", "Fee", 1, 20000, "[NOTE] Price depends on speaker – Nazaha rep or ethics expert"),
        ("Pledge Wall & Integrity Display", "Package", 1, 8000, "Interactive pledge wall + signing ceremony + LED display"),
        ("Awareness Materials & Booklets", "Per person", 200, 30, "Brochures, ethics guide booklets, policy summary"),
        ("Light Refreshments", "Per person", 200, 50, "Coffee, premium snacks"),
        ("Photography", "Half day", 1, 3500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Signage, materials, social media kit, digital assets"),
    ]),

    ("EVENT 38: Transformation Activities (أنشطة ومبادرات التحول)", "Q4 TBD", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Transformation Office Booth Setup", "Package", 1, 15000, "Interactive displays, digital transformation showcases"),
        ("Expert Speakers / Panels", "Fee", 2, 12000, "Transformation & digital innovation experts"),
        ("Interactive Workshops", "Package", 1, 12000, "Hands-on transformation activities, design thinking"),
        ("Awareness & Promotional Materials", "Per person", 200, 30, "Branded booklets, digital roadmap guides"),
        ("Light Refreshments", "Per person", 200, 45, "Coffee, snacks, lunch boxes"),
        ("Photography & Videography", "Package", 1, 8000, "Event documentation + highlight video"),
        ("Event Branding & Design", "Package", 1, 14000, "Transformation-themed signage, materials, social media kit"),
    ]),

    ("EVENT 39: Ideal Office Competition (المكتب المثالي)", "Q4 TBD", 200, "SERA HQ", [
        ("Venue", "SERA HQ", 1, 0, "Internal – no rental cost"),
        ("Evaluation Program Setup", "Package", 1, 10000, "Criteria framework, scoring system, inspections"),
        ("Award Trophies & Certificates", "Set", 1, 12000, "Best office trophies + runner-up + certificates"),
        ("Inspection & Judging Panel", "Package", 1, 8000, "External judges, evaluation sheets, site visits"),
        ("Ceremony Setup (Mini stage)", "Package", 1, 10000, "Award stage, backdrop, audio, podium"),
        ("Light Refreshments", "Per person", 200, 40, "Coffee, snacks for ceremony"),
        ("Photography", "Half day", 1, 3500, "Award ceremony documentation"),
        ("Event Branding & Design", "Package", 1, 14000, "Competition signage, certificates, social media kit"),
    ]),

    ("EVENT 40: Arabic Language Day", "December 18", 200, "Cultural Venue", [
        ("Venue (Cultural setting)", "3 hours", 1, 20000, "Cultural center or elegant hall"),
        ("Poet / Arabic Author Guest", "Fee", 1, 15000, "Arabic literature figure + book signing"),
        ("Live Calligraphy Workshop", "Package", 1, 10000, "Professional calligrapher + materials for 80"),
        ("Arabic Poetry Reading / Recital", "Package", 1, 9000, "Professional recital performance + music"),
        ("Traditional Arabic Refreshments", "Per person", 200, 60, "Arabic coffee, dates, premium sweets, luqaimat"),
        ("Cultural Décor (Calligraphy art)", "Package", 1, 8000, "Arabic calligraphy displays + art + manuscripts"),
        ("Photography", "3 hours", 1, 4500, "Event coverage"),
        ("Event Branding & Design", "Package", 1, 14000, "Arabic-themed invitations, signage, social media kit, digital assets"),
    ]),

    ("EVENT 41: Year-End Celebration", "Late December", 300, "5-Star Hotel – Premium Gala", [
        ("Venue Rental (Premium Gala setup)", "Evening", 1, 70000, "[NOTE] Book 3 months ahead – December peak season"),
        ("Catering – Premium Gala Dinner", "Per person", 300, 395, "Live cooking stations, international premium menu + dessert bar"),
        ("Stage & Premium Year-End Décor", "Package", 1, 48000, "Elegant theme + year-end branding + immersive lighting design"),
        ("Premium Entertainment (Band + Acts)", "Package", 1, 38000, "High-quality headline entertainment + performer"),
        ("LED Video Wall & Full AV Production", "Package", 1, 26000, "Year-review video playback + full production + effects"),
        ("Annual Awards Ceremony Package", "Package", 1, 20000, "15 trophies, recognition displays, awards video, frames"),
        ("Photography & Videography (Full team)", "Package", 1, 22000, "Complete professional coverage + same-day preview"),
        ("Year-End Premium Gifts", "Per person", 300, 185, "Premium thank-you gift box with personalized item"),
        ("Professional MC (Premium)", "Full event", 1, 15000, "High-profile gala host"),
        ("Event Branding & Design", "Package", 1, 18000, "Invitations, backdrop, table cards, menus, signage, social media, video"),
    ]),

    # ═══ NEWSLETTERS (52 weekly issues) ═══
    ("NEWSLETTERS", None, None, None, None),

    ("ANNUAL: 52 Weekly Internal Newsletters", "Full Year 2026", 52, "Digital + Print Distribution", [
        ("Newsletter Content Writing (Arabic)", "Per issue", 52, 2500, "Professional Arabic content writer, research, interviews"),
        ("Newsletter Content Writing (English)", "Per issue", 52, 2000, "English version / bilingual adaptation"),
        ("Graphic Design & Layout", "Per issue", 52, 2200, "Custom branded template, infographics, illustrations"),
        ("Photography for Newsletter Content", "Per month", 12, 3000, "Monthly photoshoot for newsletter imagery"),
        ("Editorial Management & Review", "Per month", 12, 4500, "Editor-in-chief, editorial calendar, quality control"),
        ("Digital Distribution Platform", "Per month", 12, 1500, "Email platform, analytics, distribution management"),
        ("Print Production (Limited VIP copies)", "Per issue", 52, 800, "[NOTE] 50 printed copies per issue for VIP/leadership distribution"),
        ("Annual Newsletter Strategy & Planning", "Package", 1, 15000, "Content strategy, editorial calendar, themes alignment"),
    ]),
]

# ── Write financial data ──
# Track event subtotal rows for quarter formulas
event_subtotal_rows = []
quarter_labels = {}
current_quarter = None
q_start_rows = {}
q_event_subtotals = {"Q1": [], "Q2": [], "Q3": [], "Q4": [], "NEWSLETTERS": []}

for event_data in events:
    name = event_data[0]

    # Quarter marker
    if name in ("Q1", "Q2", "Q3", "Q4", "NEWSLETTERS"):
        current_quarter = name
        r += 1
        add_title(ws3, r, f"{name} EVENTS", 6)
        r += 1
        q_start_rows[current_quarter] = r
        continue

    date_str = event_data[1]
    attendance = event_data[2]
    venue_type = event_data[3]
    items = event_data[4]

    # Event header
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws3.cell(row=r, column=1, value=f"{name} | Date: {date_str} | Attendance: {attendance} | {venue_type}")
    cell.font = Font(name='Arial', size=11, bold=True, color='1B4332')
    cell.fill = PatternFill(start_color='B7E4C7', end_color='B7E4C7', fill_type='solid')
    r += 1

    # Column headers
    headers = ["Cost Item", "Unit", "Qty", "Unit Price (SAR)", "Total (SAR)", "Notes"]
    for ci, h in enumerate(headers, 1):
        ws3.cell(row=r, column=ci, value=h)
    style_header_row(ws3, r, 6)
    r += 1

    # Line items
    item_start_row = r
    for idx, (desc, unit, qty, price, note) in enumerate(items):
        ws3.cell(row=r, column=1, value=desc).font = normal_font
        ws3.cell(row=r, column=2, value=unit).font = normal_font
        ws3.cell(row=r, column=3, value=qty).font = normal_font
        ws3.cell(row=r, column=3).alignment = center_align
        ws3.cell(row=r, column=4, value=price).font = normal_font
        ws3.cell(row=r, column=4).number_format = sar_format
        # FORMULA: Total = Qty × Unit Price
        ws3.cell(row=r, column=5).value = f"=C{r}*D{r}"
        ws3.cell(row=r, column=5).font = sub_font
        ws3.cell(row=r, column=5).number_format = sar_format
        ws3.cell(row=r, column=6, value=note).font = note_font if "[NOTE]" in (note or "") else Font(name='Arial', size=9, color='666666')

        for c in range(1, 7):
            ws3.cell(row=r, column=c).border = thin_border
            ws3.cell(row=r, column=c).alignment = left_align if c in (1,6) else center_align
            if idx % 2 == 1:
                ws3.cell(row=r, column=c).fill = alt_fill
        r += 1

    item_end_row = r - 1

    # Pre-commission subtotal
    ws3.cell(row=r, column=1, value="Subtotal (Before Commission)").font = sub_font
    ws3.cell(row=r, column=5).value = f"=SUM(E{item_start_row}:E{item_end_row})"
    ws3.cell(row=r, column=5).font = sub_font
    ws3.cell(row=r, column=5).number_format = sar_format
    for c in range(1, 7):
        ws3.cell(row=r, column=c).border = thin_border
    pre_commission_row = r
    r += 1

    # Agency Commission (15%) - FORMULA
    ws3.cell(row=r, column=1, value="Agency Commission (15%)").font = sub_font
    ws3.cell(row=r, column=2, value="15%").font = normal_font
    ws3.cell(row=r, column=5).value = f"=E{pre_commission_row}*0.15"
    ws3.cell(row=r, column=5).font = sub_font
    ws3.cell(row=r, column=5).number_format = sar_format
    ws3.cell(row=r, column=6, value="Agency management, planning & execution fee").font = Font(name='Arial', size=9, color='666666')
    for c in range(1, 7):
        ws3.cell(row=r, column=c).border = thin_border
    commission_row = r
    r += 1

    # EVENT SUBTOTAL = Pre-commission + Commission
    ws3.cell(row=r, column=1, value=f"▶ {name} – TOTAL").font = total_font
    ws3.cell(row=r, column=5).value = f"=E{pre_commission_row}+E{commission_row}"
    ws3.cell(row=r, column=5).font = total_font
    ws3.cell(row=r, column=5).number_format = sar_format
    style_subtotal_row(ws3, r, 6)
    event_subtotal_rows.append(r)
    q_event_subtotals[current_quarter].append(r)
    r += 2  # blank row

# ── QUARTER TOTALS ──
r += 1
add_title(ws3, r, "QUARTERLY & ANNUAL TOTALS", 6); r += 1

quarter_total_rows = {}
for q_name in ["Q1", "Q2", "Q3", "Q4", "NEWSLETTERS"]:
    label = f"{q_name} TOTAL" if q_name != "NEWSLETTERS" else "52 WEEKLY NEWSLETTERS TOTAL"
    ws3.cell(row=r, column=1, value=label).font = total_font
    formula_parts = [f"E{row}" for row in q_event_subtotals[q_name]]
    ws3.cell(row=r, column=5).value = f"={'+'.join(formula_parts)}"
    ws3.cell(row=r, column=5).font = total_font
    ws3.cell(row=r, column=5).number_format = sar_format
    count_label = f"{len(q_event_subtotals[q_name])} events" if q_name != "NEWSLETTERS" else "52 issues / year"
    ws3.cell(row=r, column=6, value=count_label).font = normal_font
    for c in range(1, 7):
        ws3.cell(row=r, column=c).fill = PatternFill(start_color='52B788', end_color='52B788', fill_type='solid')
        ws3.cell(row=r, column=c).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws3.cell(row=r, column=c).border = thin_border
    quarter_total_rows[q_name] = r
    r += 1

# Events Grand Total
r += 1
ws3.cell(row=r, column=1, value="ALL EVENTS + NEWSLETTERS SUBTOTAL (Excl. Sports)").font = grand_font
events_total_formula = f"=E{quarter_total_rows['Q1']}+E{quarter_total_rows['Q2']}+E{quarter_total_rows['Q3']}+E{quarter_total_rows['Q4']}+E{quarter_total_rows['NEWSLETTERS']}"
ws3.cell(row=r, column=5).value = events_total_formula
ws3.cell(row=r, column=5).font = grand_font
ws3.cell(row=r, column=5).number_format = sar_format
style_subtotal_row(ws3, r, 6, fill=grand_fill)
events_grand_total_row = r
r += 2

# Reference to sports sheet
ws3.cell(row=r, column=1, value="Sports Events Total").font = sub_font
ws3.cell(row=r, column=5).value = f"='Sports Budget'!E2"
ws3.cell(row=r, column=5).font = sub_font
ws3.cell(row=r, column=5).number_format = sar_format
ws3.cell(row=r, column=6, value="← Linked from Sports Budget sheet").font = link_font
for c in range(1, 7):
    ws3.cell(row=r, column=c).border = thin_border
sports_ref_row = r
r += 2

# GRAND TOTAL (ex-VAT)
ws3.cell(row=r, column=1, value="GRAND TOTAL (Excluding VAT)").font = grand_font
ws3.cell(row=r, column=5).value = f"=E{events_grand_total_row}+E{sports_ref_row}"
ws3.cell(row=r, column=5).font = grand_font
ws3.cell(row=r, column=5).number_format = sar_format
style_subtotal_row(ws3, r, 6, fill=grand_fill)
grand_total_row = r
r += 1

# VAT
ws3.cell(row=r, column=1, value="VAT (15%)").font = sub_font
ws3.cell(row=r, column=5).value = f"=E{grand_total_row}*0.15"
ws3.cell(row=r, column=5).font = sub_font
ws3.cell(row=r, column=5).number_format = sar_format
for c in range(1, 7):
    ws3.cell(row=r, column=c).border = thin_border
vat_row = r
r += 1

# GRAND TOTAL (inc. VAT)
ws3.cell(row=r, column=1, value="★ GRAND TOTAL (Including 15% VAT)").font = grand_font
ws3.cell(row=r, column=5).value = f"=E{grand_total_row}+E{vat_row}"
ws3.cell(row=r, column=5).font = grand_font
ws3.cell(row=r, column=5).number_format = sar_format
style_subtotal_row(ws3, r, 6, fill=PatternFill(start_color='000000', end_color='000000', fill_type='solid'))
final_total_row = r
r += 2

# Contingency
ws3.cell(row=r, column=1, value="With 10% Contingency Reserve").font = sub_font
ws3.cell(row=r, column=5).value = f"=E{final_total_row}*1.1"
ws3.cell(row=r, column=5).font = sub_font
ws3.cell(row=r, column=5).number_format = sar_format
for c in range(1, 7):
    ws3.cell(row=r, column=c).border = thin_border
    ws3.cell(row=r, column=c).fill = note_fill


# ════════════════════════════════════════════════════════════
# SHEET 4: SPORTS BUDGET
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Sports Budget")
ws4.sheet_properties.tabColor = "2D6A4F"

ws4.column_dimensions['A'].width = 42
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 8
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 50

# Row 1: Title, Row 2: sports grand total placeholder (will be formula)
r4 = 1
add_title(ws4, r4, "SERA 2026 — SPORTS EVENTS BUDGET", 6); r4 += 1
# Reserve row 2 for SPORTS GRAND TOTAL (referenced by Financial sheet)
ws4.cell(row=2, column=1, value="SPORTS GRAND TOTAL (inc. 15% Commission)").font = grand_font
# We'll fill the formula after writing data
sports_grand_total_row_ref = 2
for c in range(1, 7):
    ws4.cell(row=2, column=c).fill = grand_fill
    ws4.cell(row=2, column=c).font = grand_font
    ws4.cell(row=2, column=c).border = thin_border
r4 = 4

# Football
ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=6)
ws4.cell(row=r4, column=1, value="FOOTBALL TOURNAMENT | 10 Teams x 15 Players = 150 Players | Men Only | 3 Weekends (6 match days)").font = Font(name='Arial', size=11, bold=True, color='1B4332')
ws4.cell(row=r4, column=1).fill = PatternFill(start_color='B7E4C7', end_color='B7E4C7', fill_type='solid')
r4 += 1

headers = ["Cost Item", "Unit", "Qty", "Unit Price (SAR)", "Total (SAR)", "Notes"]
for ci, h in enumerate(headers, 1):
    ws4.cell(row=r4, column=ci, value=h)
style_header_row(ws4, r4, 6)
r4 += 1

football_items = [
    ("Football Pitch Rental (2 pitches)", "Per day", 6, 7000, "6 match days, 2 parallel pitches"),
    ("Team Jerseys (Numbered + SERA logo)", "Per set of 15", 10, 2800, "10 teams x 15 jerseys, unique colors"),
    ("Tournament Match Balls (Official)", "Per ball", 15, 180, "Official match balls + spares"),
    ("Training Equipment (Cones, bibs, etc.)", "Package", 1, 2000, "Shared training equipment"),
    ("Main Referees", "Per match", 15, 500, "[NOTE] Riyadh-based Saudi FA certified – no travel"),
    ("Assistant Referees (Linesmen)", "Per match", 15, 300, "2 per match"),
    ("Tournament Coordinator", "Per tournament", 1, 8000, "Full tournament management + bracket"),
    ("First Aid / Medical Staff", "Per day", 6, 1000, "On-site medical support – mandatory"),
    ("Champion Trophy (Large Crystal)", "Trophy", 1, 4500, "Premium crystal trophy"),
    ("Runner-Up Trophy", "Trophy", 1, 2500, ""),
    ("Third Place Trophy", "Trophy", 1, 2000, ""),
    ("MVP Award", "Award", 1, 600, ""),
    ("Top Scorer Award (Golden Boot)", "Award", 1, 600, ""),
    ("Best Goalkeeper Award", "Award", 1, 600, ""),
    ("Best Player per Team Award", "Award", 10, 600, "[NOTE] Optional – 1 per team recognition"),
    ("Medals (Top 3 teams)", "Per medal", 45, 60, "15 players x 3 teams"),
    ("Player Refreshments", "Per day", 6, 2000, "Water, isotonic drinks, fruits, snacks"),
    ("Post-Match Light Meals", "Per day", 6, 1200, "[NOTE] Optional – post-match meal for players"),
    ("Sports Photographer", "Per day", 6, 2000, "Match + ceremony photos"),
    ("Digital Scoreboard Display", "Per day", 6, 600, "[NOTE] Digital display if venue supports"),
    ("Social Media Tournament Coverage", "Package", 1, 3000, "Live updates, stories, highlights"),
    ("Opening Ceremony", "Package", 1, 12000, "Stage, national anthem, first-kick ceremony"),
]

fb_start = r4
for idx, (desc, unit, qty, price, note) in enumerate(football_items):
    ws4.cell(row=r4, column=1, value=desc).font = normal_font
    ws4.cell(row=r4, column=2, value=unit).font = normal_font
    ws4.cell(row=r4, column=3, value=qty).font = normal_font
    ws4.cell(row=r4, column=4, value=price).font = normal_font
    ws4.cell(row=r4, column=4).number_format = sar_format
    ws4.cell(row=r4, column=5).value = f"=C{r4}*D{r4}"
    ws4.cell(row=r4, column=5).number_format = sar_format
    ws4.cell(row=r4, column=6, value=note).font = note_font if "[NOTE]" in (note or "") else Font(name='Arial', size=9, color='666666')
    for c in range(1, 7):
        ws4.cell(row=r4, column=c).border = thin_border
        if idx % 2 == 1:
            ws4.cell(row=r4, column=c).fill = alt_fill
    r4 += 1
fb_end = r4 - 1

# Football subtotal
ws4.cell(row=r4, column=1, value="Football – Subtotal (Before Commission)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=SUM(E{fb_start}:E{fb_end})"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
fb_subtotal_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="Agency Commission (15%)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=E{fb_subtotal_row}*0.15"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
fb_commission_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="▶ FOOTBALL TOURNAMENT – TOTAL").font = total_font
ws4.cell(row=r4, column=5).value = f"=E{fb_subtotal_row}+E{fb_commission_row}"
ws4.cell(row=r4, column=5).font = total_font
ws4.cell(row=r4, column=5).number_format = sar_format
style_subtotal_row(ws4, r4, 6)
fb_total_row = r4
r4 += 3

# Padel
ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=6)
ws4.cell(row=r4, column=1, value="PADEL TOURNAMENT | 10 Teams (Doubles) = 20 Players | Men Only | 1-2 Weekends").font = Font(name='Arial', size=11, bold=True, color='1B4332')
ws4.cell(row=r4, column=1).fill = PatternFill(start_color='B7E4C7', end_color='B7E4C7', fill_type='solid')
r4 += 1

for ci, h in enumerate(headers, 1):
    ws4.cell(row=r4, column=ci, value=h)
style_header_row(ws4, r4, 6)
r4 += 1

padel_items = [
    ("Padel Court Rental", "Per 90-min slot", 25, 480, "25 matches, [NOTE] Group booking discount may apply"),
    ("Tournament Padel Balls", "Dozen", 6, 200, "Fresh balls per round"),
    ("Branded Player Towels", "Per player", 20, 50, "Tournament souvenir"),
    ("Branded Water Bottles (Reusable)", "Per player", 20, 40, "Branded reusable"),
    ("Referees / Court Coordinators", "Per day", 4, 700, "[NOTE] Riyadh-based – no travel"),
    ("Tournament Coordinator", "Package", 1, 5000, "Bracket management + logistics"),
    ("Champion Trophy (Team)", "Trophy", 1, 3000, ""),
    ("Runner-Up Trophy", "Trophy", 1, 2000, ""),
    ("MVP Award", "Award", 1, 500, ""),
    ("Best Match Award", "Award", 1, 500, ""),
    ("Fair Play Award", "Award", 1, 500, ""),
    ("Medals (Top 2 teams)", "Per medal", 4, 60, "2 players x 2 teams"),
    ("Player Refreshments", "Per day", 4, 1000, "Water, isotonic, fruits"),
    ("Light Snacks & Post-Match Food", "Per day", 4, 700, "Post-match refreshments"),
    ("Sports Photographer", "Per day", 4, 1500, "Tournament coverage"),
    ("Social Media Coverage", "Package", 1, 2000, "Live updates"),
    ("Awards Ceremony (Combined with Football)", "Package", 1, 15000, "Venue, dinner, trophy presentation – combined"),
]

pd_start = r4
for idx, (desc, unit, qty, price, note) in enumerate(padel_items):
    ws4.cell(row=r4, column=1, value=desc).font = normal_font
    ws4.cell(row=r4, column=2, value=unit).font = normal_font
    ws4.cell(row=r4, column=3, value=qty).font = normal_font
    ws4.cell(row=r4, column=4, value=price).font = normal_font
    ws4.cell(row=r4, column=4).number_format = sar_format
    ws4.cell(row=r4, column=5).value = f"=C{r4}*D{r4}"
    ws4.cell(row=r4, column=5).number_format = sar_format
    ws4.cell(row=r4, column=6, value=note).font = note_font if "[NOTE]" in (note or "") else Font(name='Arial', size=9, color='666666')
    for c in range(1, 7):
        ws4.cell(row=r4, column=c).border = thin_border
        if idx % 2 == 1:
            ws4.cell(row=r4, column=c).fill = alt_fill
    r4 += 1
pd_end = r4 - 1

ws4.cell(row=r4, column=1, value="Padel – Subtotal (Before Commission)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=SUM(E{pd_start}:E{pd_end})"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
pd_subtotal_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="Agency Commission (15%)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=E{pd_subtotal_row}*0.15"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
pd_commission_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="▶ PADEL TOURNAMENT – TOTAL").font = total_font
ws4.cell(row=r4, column=5).value = f"=E{pd_subtotal_row}+E{pd_commission_row}"
ws4.cell(row=r4, column=5).font = total_font
ws4.cell(row=r4, column=5).number_format = sar_format
style_subtotal_row(ws4, r4, 6)
pd_total_row = r4

r4 += 3

# Walking Challenge & SERA Quarterly Friday
ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=6)
ws4.cell(row=r4, column=1, value="WALKING CHALLENGE & SERA QUARTERLY FRIDAY (جمعة سيرا) | 150-200 Participants | Annual Challenge + 4 Quarterly Events").font = Font(name='Arial', size=11, bold=True, color='1B4332')
ws4.cell(row=r4, column=1).fill = PatternFill(start_color='B7E4C7', end_color='B7E4C7', fill_type='solid')
r4 += 1

for ci, h in enumerate(headers, 1):
    ws4.cell(row=r4, column=ci, value=h)
style_header_row(ws4, r4, 6)
r4 += 1

walking_items = [
    ("Fitness Tracking App / Platform License", "Annual", 1, 15000, "Step tracking, leaderboards, team challenges"),
    ("Branded Sports Gear (T-shirts, bands)", "Per person", 200, 60, "Branded walking challenge kit"),
    ("Challenge Management & Coordination", "Annual", 1, 8000, "Full challenge administration + results"),
    ("Quarterly Walking Route Planning & Permits", "Per event", 4, 3000, "Route setup, permissions, marshals"),
    ("Hydration Stations (Quarterly events)", "Per event", 4, 4000, "Water, isotonic drinks, fruits along route"),
    ("First Aid Support (Quarterly events)", "Per event", 4, 3000, "On-route medical support – mandatory"),
    ("Participant Kits (Quarterly events)", "Per person", 600, 40, "[NOTE] ~150 participants x 4 quarterly events"),
    ("Photography (Quarterly events)", "Per event", 4, 3000, "Walking event coverage"),
    ("Medals & Certificates (Quarterly)", "Per event", 4, 5000, "Top finishers medals + participation certificates"),
    ("Annual Challenge Grand Prizes", "Package", 1, 15000, "Top 10 walkers prizes + team prizes"),
    ("Social Media Coverage", "Package", 1, 3000, "Challenge updates, leaderboard posts, highlights"),
    ("Friday Gathering Refreshments (Quarterly)", "Per event", 4, 3500, "[NOTE] SERA Quarterly Friday gathering refreshments"),
]

wk_start = r4
for idx, (desc, unit, qty, price, note) in enumerate(walking_items):
    ws4.cell(row=r4, column=1, value=desc).font = normal_font
    ws4.cell(row=r4, column=2, value=unit).font = normal_font
    ws4.cell(row=r4, column=3, value=qty).font = normal_font
    ws4.cell(row=r4, column=4, value=price).font = normal_font
    ws4.cell(row=r4, column=4).number_format = sar_format
    ws4.cell(row=r4, column=5).value = f"=C{r4}*D{r4}"
    ws4.cell(row=r4, column=5).number_format = sar_format
    ws4.cell(row=r4, column=6, value=note).font = note_font if "[NOTE]" in (note or "") else Font(name='Arial', size=9, color='666666')
    for c in range(1, 7):
        ws4.cell(row=r4, column=c).border = thin_border
        if idx % 2 == 1:
            ws4.cell(row=r4, column=c).fill = alt_fill
    r4 += 1
wk_end = r4 - 1

ws4.cell(row=r4, column=1, value="Walking & Friday Events – Subtotal (Before Commission)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=SUM(E{wk_start}:E{wk_end})"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
wk_subtotal_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="Agency Commission (15%)").font = sub_font
ws4.cell(row=r4, column=5).value = f"=E{wk_subtotal_row}*0.15"
ws4.cell(row=r4, column=5).font = sub_font
ws4.cell(row=r4, column=5).number_format = sar_format
for c in range(1, 7):
    ws4.cell(row=r4, column=c).border = thin_border
wk_commission_row = r4
r4 += 1

ws4.cell(row=r4, column=1, value="▶ WALKING & FRIDAY EVENTS – TOTAL").font = total_font
ws4.cell(row=r4, column=5).value = f"=E{wk_subtotal_row}+E{wk_commission_row}"
ws4.cell(row=r4, column=5).font = total_font
ws4.cell(row=r4, column=5).number_format = sar_format
style_subtotal_row(ws4, r4, 6)
wk_total_row = r4

# Now fill the sports grand total (row 2)
ws4.cell(row=2, column=5).value = f"=E{fb_total_row}+E{pd_total_row}+E{wk_total_row}"
ws4.cell(row=2, column=5).number_format = sar_format


# ════════════════════════════════════════════════════════════
# SHEET 1: EVENT CALENDAR OVERVIEW
# ════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Event Calendar")
ws1.sheet_properties.tabColor = "40916C"
wb.move_sheet(ws1, offset=-2)  # Move to first position

ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 8
ws1.column_dimensions['F'].width = 22
ws1.column_dimensions['G'].width = 16
ws1.column_dimensions['H'].width = 12

r1 = 1
add_title(ws1, r1, "SERA 2026 — ANNUAL EVENT CALENDAR & BUDGET OVERVIEW", 8); r1 += 1
ws1.merge_cells(start_row=r1, start_column=1, end_row=r1, end_column=8)
ws1.cell(row=r1, column=1, value="Saudi Electricity Regulatory Authority (هيئة تنظيم الكهرباء) | All Events in Riyadh | Agency Commission: 15%").font = Font(name='Arial', size=10, italic=True, color='1B4332')
r1 += 2

cal_headers = ["#", "Event Name", "Arabic Name", "Date", "Qtr", "Category", "Est. Attendance", "Event Tier"]
for ci, h in enumerate(cal_headers, 1):
    ws1.cell(row=r1, column=ci, value=h)
style_header_row(ws1, r1, 8)
r1 += 1

cal_events = [
    (1, "Annual Meeting 2026", "الاجتماع السنوي", "February 3", "Q1", "Corporate Conference", 300, "Major"),
    (2, "Founding Day", "يوم التأسيس", "February 22", "Q1", "National Celebration", 300, "Light"),
    (3, "Employee Appreciation Day", "يوم تقدير الموظف", "March 2", "Q1", "Recognition Gala", 300, "Major"),
    (4, "International Women's Day", "اليوم العالمي للمرأة", "March 8", "Q1", "Awareness/Celebration", 150, "Medium"),
    (5, "Saudi Flag Day", "يوم العلم السعودي", "March 11", "Q1", "Patriotic Celebration", 300, "Light"),
    (6, "Ramadan Iftar", "إفطار رمضان", "Mid-March", "Q1", "Corporate Iftar", 300, "Major"),
    (7, "Mother's Day", "يوم الأم", "March 21", "Q1", "Appreciation Event", 130, "Medium"),
    (8, "Saudi Green Initiative", "المبادرة الخضراء السعودية", "March 27", "Q1", "Sustainability Awareness", 250, "Light"),
    (9, "Ehsan Philanthropy Campaign", "حملة إحسان", "March", "Q1", "CSR/Charity", 200, "Light"),
    (10, "Eid Al-Fitr Celebration", "احتفال عيد الفطر", "Early April", "Q2", "Major Gala", 300, "Major"),
    (11, "Creativity & Innovation Day", "يوم الإبداع والابتكار", "April 21", "Q2", "Professional Development", 200, "Medium"),
    (12, "International Tea Day", "اليوم العالمي للشاي", "May 21", "Q2", "Social/Wellness", 250, "Light"),
    (13, "Joy Clothing Drive", "كسوة فرح", "May", "Q2", "CSR/Charity", 200, "Light"),
    (14, "World No Tobacco Day", "اليوم العالمي لمكافحة التدخين", "May 31", "Q2", "Health Awareness", 200, "Light"),
    (15, "Blood Donation Day", "يوم التبرع بالدم", "June 14", "Q2", "CSR/Health", 120, "Light"),
    (16, "Father's Day", "يوم الأب", "June 15", "Q2", "Appreciation Event", 160, "Medium"),
    (17, "Eid Al-Adha Celebration", "احتفال عيد الأضحى", "Mid-June", "Q2", "Major Gala", 300, "Major"),
    (18, "Summer SERA (صيف سيرا)", "صيف سيرا", "July 20", "Q3", "Seasonal Family Event", 500, "Major"),
    (19, "SERA Winter (شتوية سيرا)", "شتوية سيرا", "August 19", "Q3", "Seasonal Indoor Event", 300, "Medium"),
    (20, "World First Aid Day", "اليوم العالمي للإسعافات الأولية", "September 13", "Q3", "Health Awareness", 200, "Light"),
    (21, "تزود Skills Development", "تزوّد", "September", "Q3", "Professional Development", 200, "Medium"),
    (22, "World Alzheimer's Day", "اليوم العالمي للزهايمر", "September 21", "Q3", "Health Awareness", 200, "Light"),
    (23, "وش دورنا – What's Our Role", "وش دورنا", "September", "Q3", "Internal Engagement", 200, "Light"),
    (24, "Saudi National Day (96th)", "اليوم الوطني السعودي", "September 23", "Q3", "National Celebration", 300, "Major"),
    (25, "Environment Week", "أسبوع البيئة", "Q3 TBD", "Q3", "Sustainability Awareness", 250, "Medium"),
    (26, "International Coffee Day", "اليوم العالمي للقهوة", "October 1", "Q4", "Social/Wellness", 250, "Light"),
    (27, "Breast Cancer Awareness Day", "اليوم العالمي لسرطان الثدي", "October", "Q4", "Health Awareness", 200, "Light"),
    (28, "Cybersecurity Exhibition", "معرض الأمن السيبراني", "October TBD", "Q4", "Professional/Awareness", 250, "Medium"),
    (29, "World Mental Health Day", "اليوم العالمي للصحة النفسية", "October 10", "Q4", "Health Awareness", 200, "Light"),
    (30, "World Savings Day", "اليوم العالمي للادخار", "October 31", "Q4", "Financial Awareness", 200, "Light"),
    (31, "Quality Day", "يوم الجودة", "November 10", "Q4", "Professional Development", 180, "Medium"),
    (32, "Flu Vaccination Campaign", "حملة التطعيم ضد الإنفلونزا", "November", "Q4", "Health Campaign", 250, "Light"),
    (33, "World Diabetes Day", "اليوم العالمي للسكري", "November 14", "Q4", "Health Awareness", 200, "Light"),
    (34, "International Men's Day", "اليوم العالمي للرجل", "November 19", "Q4", "Appreciation Event", 200, "Medium"),
    (35, "Universal Children's Day", "اليوم العالمي للطفل", "November 20", "Q4", "CSR/Family", 200, "Medium"),
    (36, "Volunteer Day", "يوم التطوع", "December 5", "Q4", "CSR Activity", 80, "Light"),
    (37, "Anti-Corruption Day", "اليوم العالمي لمكافحة الفساد", "December 9", "Q4", "Awareness", 200, "Light"),
    (38, "Transformation Activities", "أنشطة ومبادرات التحول", "Q4 TBD", "Q4", "Digital Transformation", 200, "Medium"),
    (39, "Ideal Office Competition", "المكتب المثالي", "Q4 TBD", "Q4", "Internal Competition", 200, "Light"),
    (40, "Arabic Language Day", "اليوم العالمي للغة العربية", "December 18", "Q4", "Cultural Celebration", 200, "Medium"),
    (41, "Year-End Celebration", "حفل نهاية العام", "Late December", "Q4", "Major Gala", 300, "Major"),
    ("S1", "Football Tournament", "بطولة كرة القدم", "TBD (3 weekends)", "Sports", "Sports Competition", 150, "Major"),
    ("S2", "Padel Tournament", "بطولة البادل", "TBD (2 weekends)", "Sports", "Sports Competition", 20, "Medium"),
    ("S3", "Walking Challenge & SERA Friday", "تحدي المشي + جمعة سيرا", "Quarterly + Annual", "Sports", "Sports/Wellness", 200, "Medium"),
]

for idx, ev in enumerate(cal_events):
    for ci, val in enumerate(ev, 1):
        ws1.cell(row=r1, column=ci, value=val)
    style_data_row(ws1, r1, 8, is_alt=(idx % 2 == 1))
    # Color tier
    tier = ev[7]
    tier_cell = ws1.cell(row=r1, column=8)
    if tier == "Major":
        tier_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        tier_cell.font = Font(name='Arial', size=10, bold=True)
    elif tier == "Medium":
        tier_cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    r1 += 1

# Budget summary on calendar sheet
r1 += 2
add_section(ws1, r1, "BUDGET SUMMARY (Linked from Financial Proposal & Sports Budget)", 8)
r1 += 1
sum_headers = ["Category", "# Events", "Total (SAR)", "% of Annual", "", "", "", ""]
for ci, h in enumerate(sum_headers, 1):
    ws1.cell(row=r1, column=ci, value=h)
style_header_row(ws1, r1, 4)
r1 += 1

# Linked formulas to Financial Proposal
for q, count in [("Q1", 9), ("Q2", 8), ("Q3", 8), ("Q4", 16)]:
    ws1.cell(row=r1, column=1, value=f"{q} Events").font = sub_font
    ws1.cell(row=r1, column=2, value=count).font = normal_font
    ws1.cell(row=r1, column=3).value = f"='Financial Proposal'!E{quarter_total_rows[q]}"
    ws1.cell(row=r1, column=3).font = sub_font
    ws1.cell(row=r1, column=3).number_format = sar_format
    ws1.cell(row=r1, column=4).value = f"=C{r1}/C{r1+5}" if q != "Q4" else f"=C{r1}/C{r1+2}"
    ws1.cell(row=r1, column=4).number_format = '0%'
    for c in range(1, 5):
        ws1.cell(row=r1, column=c).border = thin_border
    r1 += 1

# Sports total
ws1.cell(row=r1, column=1, value="Sports Events").font = sub_font
ws1.cell(row=r1, column=2, value=3).font = normal_font
ws1.cell(row=r1, column=3).value = f"='Sports Budget'!E2"
ws1.cell(row=r1, column=3).font = sub_font
ws1.cell(row=r1, column=3).number_format = sar_format
for c in range(1, 5):
    ws1.cell(row=r1, column=c).border = thin_border
r1 += 1

# Grand totals on overview
ws1.cell(row=r1, column=1, value="GRAND TOTAL (ex-VAT)").font = total_font
ws1.cell(row=r1, column=3).value = f"='Financial Proposal'!E{grand_total_row}"
ws1.cell(row=r1, column=3).font = total_font
ws1.cell(row=r1, column=3).number_format = sar_format
for c in range(1, 5):
    ws1.cell(row=r1, column=c).fill = subtotal_fill
    ws1.cell(row=r1, column=c).font = total_font
    ws1.cell(row=r1, column=c).border = thin_border
r1 += 1

ws1.cell(row=r1, column=1, value="VAT (15%)").font = sub_font
ws1.cell(row=r1, column=3).value = f"=C{r1-1}*0.15"
ws1.cell(row=r1, column=3).number_format = sar_format
for c in range(1, 5):
    ws1.cell(row=r1, column=c).border = thin_border
r1 += 1

ws1.cell(row=r1, column=1, value="★ GRAND TOTAL (inc. VAT)").font = grand_font
ws1.cell(row=r1, column=3).value = f"=C{r1-2}+C{r1-1}"
ws1.cell(row=r1, column=3).font = grand_font
ws1.cell(row=r1, column=3).number_format = sar_format
for c in range(1, 5):
    ws1.cell(row=r1, column=c).fill = grand_fill
    ws1.cell(row=r1, column=c).font = grand_font
    ws1.cell(row=r1, column=c).border = thin_border


# ════════════════════════════════════════════════════════════
# SHEET 2: TECHNICAL PROPOSAL
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Technical Proposal")
ws2.sheet_properties.tabColor = "52B788"
wb.move_sheet(ws2, offset=-2)  # Move to second position

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 35
ws2.column_dimensions['F'].width = 35
ws2.column_dimensions['G'].width = 50

r2 = 1
add_title(ws2, r2, "SERA 2026 — TECHNICAL PROPOSAL (EVENT SPECIFICATIONS)", 7); r2 += 2

tech_headers = ["#", "Event Name", "Date", "Venue Recommendation", "Stage & Décor", "Audio-Visual Requirements", "Key Services & Scope"]
for ci, h in enumerate(tech_headers, 1):
    ws2.cell(row=r2, column=ci, value=h)
style_header_row(ws2, r2, 7)
r2 += 1

tech_data = [
    (1, "Annual Meeting 2026", "Feb 3", "5-Star Hotel Ballroom (Ritz-Carlton / Four Seasons)", "Stage 8m×5m, branded backdrop, podium with SERA logo", "LED Wall P2.5 5m×3m, Line-array sound, 6 wireless mics, live streaming", "Simultaneous AR/EN translation, Prof. MC, Registration QR system, Photography (2), Videography (2), Printed materials"),
    (2, "Founding Day", "Feb 22", "SERA HQ – Internal", "Heritage displays, Founding Day national theme", "Sound for anthem + speeches", "Heritage displays, commemorative items, social media live coverage"),
    (3, "Employee Appreciation Day", "Mar 2", "5-Star Hotel – Gala round-table setup", "Gala stage, premium décor, uplighting, award pedestals", "Full AV, spotlight for awards, background music system", "12 crystal trophies, Live band (2.5hr), Gift boxes, Prof. MC, Full photo/video team"),
    (4, "International Women's Day", "Mar 8", "Boutique venue or hotel ladies' section", "Elegant feminine décor, premium floral arrangements", "Projector, wireless mic, ambient music", "Female keynote + panel (3 speakers), Photo booth, Curated gift bags, Female photographer"),
    (5, "Saudi Flag Day", "Mar 11", "SERA HQ – Lobby & outdoor", "Large Saudi flag (5m+), national green/white theme", "Portable sound for anthem + speeches", "Flag ceremony, Commemorative items, Social media live coverage"),
    (6, "Ramadan Iftar", "Mid-Mar", "5-Star Hotel Ramadan Tent", "Traditional Ramadan décor: lanterns, crescents, VIP majlis", "Ambient lighting, background system", "Premium Iftar buffet, Oud player, Ramadan gift boxes, Prayer area arrangement"),
    (7, "Mother's Day", "Mar 21", "Elegant restaurant private section", "Floral-themed décor, photo area", "Basic sound for music/announcements", "Flower gifts, Special mother gifts, Entertainment, Photography"),
    (8, "Saudi Green Initiative", "Mar 27", "SERA HQ – Exhibition area", "Green/sustainability themed booths, eco displays", "Presentation setup, ambient music", "Environmental expert, Green exhibits, Eco giveaways, Sustainability displays"),
    (9, "Ehsan Philanthropy Campaign", "March", "SERA HQ – Lobby area", "Ehsan platform displays, donation stations", "Basic presentation setup", "Campaign awareness booth, QR donation stations, Impact stories display"),
    (10, "Eid Al-Fitr Celebration", "Early Apr", "5-Star Hotel Grand Ballroom", "Festive Eid gold/white, large stage, entrance arch", "Full production: LED wall, concert sound, lighting design", "Ardah troupe + live music, Eid gifts, Children's corner (optional), Prof. MC"),
    (11, "Creativity & Innovation Day", "Apr 21", "Conference Center", "Main stage + 10 department innovation booths", "Conference AV, booth displays, digital voting", "Keynote speaker, Innovation competition, 7 awards, Interactive voting"),
    (12, "International Tea Day", "May 21", "SERA HQ Common Areas", "Tea-themed décor, 3 service stations", "Background music", "Arabic/English/Asian tea stations, Gourmet pairings, Tea culture displays"),
    (13, "Joy Clothing Drive", "May", "SERA HQ – Collection area", "Collection booths, sorting stations, CSR banners", "None needed", "Collection points, volunteer coordination, sorting & distribution logistics"),
    (14, "World No Tobacco Day", "May 31", "SERA HQ – Health zone", "Anti-smoking displays, health info panels", "Presentation setup", "Health expert speaker, CO breath test station, Awareness materials"),
    (15, "Blood Donation Day", "Jun 14", "SERA HQ – Medical setup area", "Awareness materials, registration area, recovery zone", "Digital screens with awareness content", "Red Crescent partnership, Pre-event campaign, Donor gifts & certificates"),
    (16, "Father's Day", "Jun 15", "Restaurant / Event space", "Celebration décor, activity zone", "Sound system", "Gifts for fathers (50% assumed), Team-building activities"),
    (17, "Eid Al-Adha Celebration", "Mid-Jun", "5-Star Hotel Grand Ballroom", "Traditional Arabian Eid theme, large stage", "Full production: LED, concert sound, lighting", "Ardah + live music, Lamb-focused feast, Eid gifts, Prof. MC"),
    (18, "Summer SERA (صيف سيرا)", "Jul 20", "Entertainment venue – EVENING ONLY (45°C heat)", "Multi-zone: main stage, kids zone, food area, games", "Concert-grade outdoor sound & lighting", "Headline entertainment, Kids zone, 10 game booths, Food festival, Raffle, Family gifts"),
    (19, "SERA Winter (شتوية سيرا)", "Aug 19", "Indoor entertainment venue (air-conditioned)", "Winter wonderland theme, snow effects, ambient lighting", "Full AV, music, effects", "Entertainment program, Food stations, Photo booth, Winter-themed activities"),
    (20, "World First Aid Day", "Sep 13", "SERA HQ – Training area", "First aid training layout, demo stations", "Presentation setup", "CPR training workshop, First aid kits, Paramedic expert, Certification"),
    (21, "تزود Skills Development", "Sep", "Conference / Training Venue", "Professional training setup, workshop layout", "Conference AV, projector, mics", "Expert trainers (2 sessions), Training materials, Certificates, Full-day program"),
    (22, "World Alzheimer's Day", "Sep 21", "SERA HQ", "Purple awareness theme, info booth", "Presentation setup", "Healthcare expert speaker, Purple ribbons, Awareness displays"),
    (23, "وش دورنا – What's Our Role", "Sep", "SERA HQ – Meeting area", "Interactive workshop layout, role displays", "Basic AV for presentations", "Professional facilitator, Interactive exercises, Role guides, Team activities"),
    (24, "Saudi National Day (96th)", "Sep 23", "Premium venue with outdoor option", "Premium green national décor, flags, national symbols", "Full AV production, LED national content, green effects", "Ardah troupe, Cultural acts, National gifts, Green lighting effects. BOOK 3-4 MONTHS AHEAD"),
    (25, "Environment Week", "Q3 TBD", "SERA HQ – Exhibition area (5 days)", "Green/eco displays, sustainability booths, plant features", "Daily presentation setup", "5-day program, Expert speakers, Tree planting, Eco workshops, Green activities"),
    (26, "International Coffee Day", "Oct 1", "SERA HQ Common Areas", "Coffee-themed décor, 3 stations", "Background music", "Professional barista, Arabic dallah ceremony, Coffee culture displays"),
    (27, "Breast Cancer Awareness Day", "Oct", "SERA HQ – Health zone", "Pink ribbon theme, awareness panels, info displays", "Presentation setup", "Healthcare expert, Screening info, Pink awareness items, Health materials"),
    (28, "Cybersecurity Exhibition", "Oct TBD", "SERA HQ / Conference Center", "6 interactive exhibition booths, demo stations", "Full conference AV, booth monitors, demo equipment", "6 interactive demos, 2 expert speakers, Phishing simulations, Branded giveaways"),
    (29, "World Mental Health Day", "Oct 10", "SERA HQ", "Wellness décor, relaxation zone", "Presentation setup, ambient music", "Psychologist speaker, Stress workshop, Relaxation zone (massage chairs), Wellness kits"),
    (30, "World Savings Day", "Oct 31", "SERA HQ – Meeting area", "Financial literacy displays, pledge wall", "Presentation setup", "Financial expert, Savings guides, Budget planners, Pledge wall activity"),
    (31, "Quality Day", "Nov 10", "Business Hotel Conference Room", "Award stage, quality showcase display", "Conference AV: screen, sound, podium", "7 quality awards, Best practices presentations, Expert speaker"),
    (32, "Flu Vaccination Campaign", "Nov", "SERA HQ – Medical area", "Health campaign setup, vaccination stations", "Digital awareness screens", "Medical staff, Vaccination logistics, Pre-campaign awareness, Participant gifts"),
    (33, "World Diabetes Day", "Nov 14", "SERA HQ – Health zone", "Blue circle theme, diabetes awareness panels", "Presentation setup", "Endocrinologist speaker, Blood sugar screening, Health materials, Blue awareness items"),
    (34, "International Men's Day", "Nov 19", "Event space", "Celebration décor, activity area", "Sound system", "Team competitions, Recognition gifts"),
    (35, "Universal Children's Day", "Nov 20", "Family entertainment venue", "Colorful child-friendly décor, safe play areas", "Sound for entertainment", "Professional children's entertainers, Game stations, Gifts for kids"),
    (36, "Volunteer Day", "Dec 5", "Community location", "Branded volunteer banner/signage", "None needed", "2 buses transport, Volunteer T-shirts, Activity supplies, Lunch boxes"),
    (37, "Anti-Corruption Day", "Dec 9", "SERA HQ", "Integrity pledge wall, information displays", "Presentation AV", "Ethics expert/Nazaha speaker, Pledge signing, Awareness brochures"),
    (38, "Transformation Activities", "Q4 TBD", "SERA HQ – Exhibition area", "Digital transformation displays, innovation showcases", "Full AV, interactive screens", "Expert panels, Transformation workshops, Design thinking activities"),
    (39, "Ideal Office Competition", "Q4 TBD", "SERA HQ – All offices", "Competition banners, evaluation signage", "Basic AV for ceremony", "Evaluation program, Judging panel, Award ceremony, Trophies"),
    (40, "Arabic Language Day", "Dec 18", "Cultural venue", "Arabic calligraphy displays, poetry corner", "Presentation + ambient Arabic music", "Poet guest, Calligraphy workshop, Poetry recital, Cultural program"),
    (41, "Year-End Celebration", "Late Dec", "5-Star Hotel – Premium Gala", "Premium stage, LED lighting design, luxury tables", "Full gala: LED wall, concert sound, designer lighting", "Premium band + acts, 15 annual awards, Year-review video, Premium gifts, Prof. MC. BOOK 3 MONTHS AHEAD"),
]

for idx, row_data in enumerate(tech_data):
    for ci, val in enumerate(row_data, 1):
        ws2.cell(row=r2, column=ci, value=val)
    style_data_row(ws2, r2, 7, is_alt=(idx % 2 == 1))
    r2 += 1


# ════════════════════════════════════════════════════════════
# SHEET 5: ASSUMPTIONS & NOTES
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Assumptions & Notes")
ws5.sheet_properties.tabColor = "95D5B2"

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 45
ws5.column_dimensions['D'].width = 40

r5 = 1
add_title(ws5, r5, "SERA 2026 — ASSUMPTIONS, NOTES & PAYMENT TERMS", 4); r5 += 2

# Assumptions
add_section(ws5, r5, "KEY ASSUMPTIONS", 4); r5 += 1
a_headers = ["#", "Assumption", "Impact if Different", "Action Required"]
for ci, h in enumerate(a_headers, 1):
    ws5.cell(row=r5, column=ci, value=h)
style_header_row(ws5, r5, 4)
r5 += 1

assumptions = [
    ("All events in Riyadh only – no travel costs", "Other cities add 15-25% for travel, accommodation, logistics", "Confirm all events are Riyadh-based"),
    ("Reception events: 250-300 attendees", "Catering, venue, gifts scale linearly", "SERA to confirm headcount 2 weeks before"),
    ("Sports: 10 teams per sport, men only", "More teams = longer tournament + more jerseys/venue", "Registration cap at 10 teams"),
    ("All referees Riyadh-based – no travel", "Out-of-city refs add SAR 2K-5K travel each", "Use local Saudi FA/Padel Federation certified"),
    ("SERA HQ available at no cost for internal events", "External venue for all = +SAR 15K-25K per awareness event", "SERA admin to confirm HQ availability"),
    ("Standard gifts (SAR 50-150 range)", "Premium/luxury gifts (SAR 200-500) double the line item", "Confirm gift tier per event category"),
    ("SERA does not own professional AV equipment", "Own AV = reduce SAR 8K-20K per event", "Audit SERA's existing AV inventory"),
    ("Standard professional entertainment", "Celebrity = SAR 50K-500K+ additional per appearance", "Confirm entertainment tier per major event"),
    ("Ramadan 2026 dates per expected Hijri calendar", "Shifting dates affect Q1/Q2 scheduling", "Confirm when official dates announced"),
    ("2026 prices with 5-10% inflation from 2025", "If contracts delayed, additional adjustment needed", "Include price escalation clause"),
    ("Agency commission: 15% on all events", "Industry range is 10-20% depending on scope", "Standard for full-service event management"),
    ("VAT at 15% (ZATCA standard)", "Rate changes affect total", "Applied to all taxable services"),
    ("Women's Day requires female photographer & ladies venue", "Mixed-gender reduces venue restrictions", "Standard for Saudi government entities"),
    ("Family events (Summer/Children's Day) include ~200 family members", "Employees-only reduces budget by ~40%", "Confirm family invitation policy"),
    ("National Day = major-tier event with premium budget", "Standard tier saves SAR 100K+", "Typically highest-profile for government"),
    ("Children's Corner at Eid = optional (families may not be invited)", "If families attend, add SAR 12K per Eid event", "SERA to confirm family invitation policy"),
]

for idx, (assumption, impact, action) in enumerate(assumptions, 1):
    ws5.cell(row=r5, column=1, value=idx).font = normal_font
    ws5.cell(row=r5, column=2, value=assumption).font = normal_font
    ws5.cell(row=r5, column=3, value=impact).font = normal_font
    ws5.cell(row=r5, column=4, value=action).font = normal_font
    for c in range(1, 5):
        ws5.cell(row=r5, column=c).border = thin_border
        ws5.cell(row=r5, column=c).alignment = left_align
        if idx % 2 == 0:
            ws5.cell(row=r5, column=c).fill = alt_fill
    r5 += 1

r5 += 2
add_section(ws5, r5, "PAYMENT TERMS (KSA Market Standard)", 4); r5 += 1
pay_headers = ["#", "Term", "Condition", "Details"]
for ci, h in enumerate(pay_headers, 1):
    ws5.cell(row=r5, column=ci, value=h)
style_header_row(ws5, r5, 4)
r5 += 1

payments = [
    ("Advance Payment", "30% upon contract signing", "Secures venue bookings and vendor commitments"),
    ("Progress Payment", "40% two weeks before each event", "Triggers final vendor confirmations and material production"),
    ("Final Payment", "30% within 14 business days post-event", "Upon delivery of photos, videos, and evaluation report"),
    ("Cancellation 30+ days", "100% refund of advance", "Full refund minus non-recoverable vendor deposits"),
    ("Cancellation 15-30 days", "50% refund", "Venue and vendor cancellation fees apply"),
    ("Cancellation <15 days", "No refund", "Full charges apply – vendors already committed"),
    ("Event Postponement", "Rescheduling at no extra charge", "Subject to venue/vendor availability within 60 days"),
    ("Price Validity", "90 days from proposal date", "Subject to vendor rate confirmations"),
    ("Annual Contract Discount", "5-8% for full annual commitment", "Negotiable based on payment terms"),
]

for idx, (term, condition, detail) in enumerate(payments, 1):
    ws5.cell(row=r5, column=1, value=idx).font = normal_font
    ws5.cell(row=r5, column=2, value=term).font = sub_font
    ws5.cell(row=r5, column=3, value=condition).font = normal_font
    ws5.cell(row=r5, column=4, value=detail).font = normal_font
    for c in range(1, 5):
        ws5.cell(row=r5, column=c).border = thin_border
        ws5.cell(row=r5, column=c).alignment = left_align
    r5 += 1

r5 += 2
add_section(ws5, r5, "COST OPTIMIZATION OPPORTUNITIES", 4); r5 += 1
opt_headers = ["#", "Opportunity", "Potential Savings (SAR)", "Recommendation"]
for ci, h in enumerate(opt_headers, 1):
    ws5.cell(row=r5, column=ci, value=h)
style_header_row(ws5, r5, 4)
r5 += 1

optimizations = [
    ("Annual vendor contracts (preferred rates)", "200,000 – 350,000", "STRONGLY RECOMMENDED – negotiate 10-15% volume discounts"),
    ("Maximize SERA HQ for awareness events", "120,000 – 180,000", "RECOMMENDED – save venue budget for major galas"),
    ("Bundled AV equipment annual rental", "50,000 – 80,000", "RECOMMENDED – consistency + lower rates"),
    ("Bundled photography/videography contract", "40,000 – 65,000", "RECOMMENDED – familiar with SERA brand"),
    ("Reduce gift tier for awareness days", "40,000 – 60,000", "RECOMMENDED – reserve premium for celebrations"),
    ("Off-peak weekday timing for awareness events", "60,000 – 100,000", "CONSIDER – may reduce attendance"),
    ("Standard entertainment for mid-tier events", "80,000 – 130,000", "SELECTIVELY APPLY – keep premium for Eid, National Day, Year-End"),
]

for idx, (opp, savings, rec) in enumerate(optimizations, 1):
    ws5.cell(row=r5, column=1, value=idx).font = normal_font
    ws5.cell(row=r5, column=2, value=opp).font = normal_font
    ws5.cell(row=r5, column=3, value=savings).font = sub_font
    ws5.cell(row=r5, column=4, value=rec).font = normal_font
    for c in range(1, 5):
        ws5.cell(row=r5, column=c).border = thin_border
        ws5.cell(row=r5, column=c).alignment = left_align
        if idx % 2 == 0:
            ws5.cell(row=r5, column=c).fill = alt_fill
    r5 += 1


# ── Freeze panes ──
ws1.freeze_panes = 'A5'
ws2.freeze_panes = 'A4'
ws3.freeze_panes = 'A5'
ws4.freeze_panes = 'A4'
ws5.freeze_panes = 'A4'

# ── Print setup ──
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.sheet_format.defaultRowHeight = 18

# ── Save ──
output_path = "/home/user/dns/SERA_2026_Proposal_Complete.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
